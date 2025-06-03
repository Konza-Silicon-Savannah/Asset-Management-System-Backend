import jwt
from datetime import datetime, timedelta
from django.conf import settings
from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from rest_framework import status
from rest_framework.decorators import action, api_view
from rest_framework.parsers import JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from .models import Asset
from .serializers import AssetSerializer
from ..utils.token import JWTAuthentication
from openpyxl import Workbook, load_workbook

# Create your views here.
@api_view(['GET'])
def asset_types(request):
    # Return objects with id and name properties
    types = [
        {'id': 1, 'name': 'Desktop Computer'},
        {'id': 2, 'name': 'Laptop'},
        {'id': 3, 'name': 'Monitor'},
        {'id': 4, 'name': 'Printer'},
        {'id': 5, 'name': 'Server'},
        {'id': 6, 'name': 'Network Equipment'},
        {'id': 7, 'name': 'Mobile Device'},
        {'id': 8, 'name': 'Software License'},
        {'id': 9, 'name': 'Furniture'},
        {'id': 10, 'name': 'Vehicle'},
    ]
    return Response(types)  

@api_view(['GET'])
def locations(request):
    locations_list = [
        {'id': 1, 'name': 'Main Office - Floor 1'},
        {'id': 2, 'name': 'Main Office - Floor 2'},
        {'id': 3, 'name': 'Main Office - Floor 3'},
        {'id': 4, 'name': 'Warehouse'},
        {'id': 5, 'name': 'Remote Office - North'},
        {'id': 6, 'name': 'Remote Office - South'},
        {'id': 7, 'name': 'Data Center'},
        {'id': 8, 'name': 'Storage Room'},
        {'id': 9, 'name': 'Conference Room A'},
        {'id': 10, 'name': 'Conference Room B'},
    ]
    return Response(locations_list)

@api_view(['GET'])
def departments(request):
    departments_list = [
        {'id': 1, 'name': 'Information Technology'},
        {'id': 2, 'name': 'Human Resources'},
        {'id': 3, 'name': 'Finance'},
        {'id': 4, 'name': 'Marketing'},
        {'id': 5, 'name': 'Sales'},
        {'id': 6, 'name': 'Operations'},
        {'id': 7, 'name': 'Administration'},
        {'id': 8, 'name': 'Research & Development'},
        {'id': 9, 'name': 'Customer Service'},
        {'id': 10, 'name': 'Legal'},
    ]
    return Response(departments_list)

@api_view(['GET'])
def status_list(request):
    status_options = [
        {'id': 1, 'name': 'new'},
        {'id': 2, 'name': 'disposal'},
        {'id': 3, 'name': 'good'},
        {'id': 4, 'name': 'damaged'},
        
    ]
    return Response(status_options)


class AssetViewSet(ModelViewSet):
    queryset = Asset.objects.all().order_by("-created_at")
    serializer_class = AssetSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        assets = self.queryset

        status = self.request.GET.get("status")
        if status and status != "all":
            assets = self.queryset.filter(status=status)

        search = self.request.GET.get("search")
        if search:
            assets = self.queryset.filter(Q(name__icontains=search) | Q(model__icontains=search) | Q(type__icontains=search))

        return assets

    @action(detail=False, methods=['get'], url_path='available')
    def available_assets(self, request):
        available = Asset.objects.exclude(status__in=['disposal', 'damaged'])

        asset_type = request.GET.get("type")
        if asset_type and asset_type != "all":
            available = available.filter(type=asset_type)

        search = request.GET.get("search")
        if search:
            available = available.filter(Q(name__icontains=search) | Q(model__icontains=search)| Q(serial_no__icontains=search) | Q(asset_tag__icontains=search))

        serializer = self.get_serializer(available, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='types')
    def available_asset_types(self, request):
        types = Asset.objects.values_list('type', flat=True).distinct()
        return Response(sorted(set(filter(None, types))), status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='excel')
    def import_data(self, request):
        file = request.FILES.get('file')
        staff = request.data.get('staff')  # Optional: if you want to track who uploaded

        if not file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(file)
            ws = wb.active
            count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming row 1 is header
                if row is None or len(row) != 9:
                    continue  # Skip malformed rows

                name, asset_tag, serial_no, model, purchase_cost, department, type, location, description = row
                exists = Asset.objects.filter(serial_no=serial_no, asset_tag=asset_tag).exists()

                if not exists:
                    Asset.objects.create(
                        name=name,
                        asset_tag=asset_tag,
                        serial_no=serial_no,
                        model=model,
                        purchase_cost=purchase_cost,
                        department=department,
                        type=type,
                        location=location,
                        description=description
                    )
                    count += 1

            if count == 0:
                return Response({"message": "All assets already exist in the database."}, status=status.HTTP_200_OK)

            return Response({"message": f"Successfully imported {count} asset(s)."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            print("Error during import:", e)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)